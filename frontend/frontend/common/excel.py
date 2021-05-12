import requests
import json
from datetime import datetime, timedelta
from frontend.common import common as mcm
from frontend.common import constant as mcs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from frontend.html2pdf.templatetags import pdfgeneration
import os.path
import random
import time

def render_excel(context, excel_filename):
    if context['contentAlias'] == 'exportorders':
        context['excelcontent'] = ordercontent(context)
        render_ordercontent_excel(context, excel_filename)


def ordercontent(context):
    try:
        request = context['request']
        params = context['param']

    except:
        params = []

    headerinfo = [
        {'field': 'username', 'fieldAlias': 'User'},
        {'field': 'service_name', 'fieldAlias': 'Service'},
        {'field': 'description', 'fieldAlias': 'Description'},
        {'field': 'price', 'fieldAlias': 'Price'},
        {'field': 'paid_price', 'fieldAlias': 'Paid Price'},
        {'field': 'service_status_name', 'fieldAlias': 'Service Status'},
        {'field': 'date_booked', 'fieldAlias': 'Date Booked'},
        {'field': 'date_created', 'fieldAlias': 'Date Created'},
    ]

    return {'orders': params, 'headers': headerinfo}


def render_ordercontent_excel(context, excel_filename):
    wb = openpyxl.Workbook()
    wb.save(excel_filename)
    csv_file = openpyxl.load_workbook(excel_filename)
    sheet = csv_file['Sheet']
    sheet.freeze_panes = "A3"

    # title
    sheet["A1"] = "Orders (" + context["export_from"] + " ~ " + context["export_to"] + ")"
    sheet["A1"].alignment = Alignment(horizontal='center')

    # filter
    index = 2
    for k in range(0, len(context['excelcontent']['headers'])):
        letter = get_excel_default_key(k)
        sheet[letter + str(index)] = context['excelcontent']['headers'][k]['fieldAlias']
        sheet[letter + str(index)].fill = PatternFill(fgColor="8C8C8C", fill_type="solid")
        sheet[letter + str(index)].alignment = Alignment(horizontal='center')

    # content
    for i in range(0, len(context['excelcontent']['orders'])):
        index += 1
        for j in range(0, len(context['excelcontent']['headers'])):
            letter = get_excel_default_key(j)
            sheet[letter + str(index)] = context['excelcontent']['orders'][i][
                context['excelcontent']['headers'][j]['field']]
            sheet[letter + str(index)].alignment = Alignment(horizontal='center')

    FullRange = "A2:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    sheet.auto_filter.ref = FullRange
    set_width_adjustment(sheet, '')

    # merge title
    letter = get_excel_default_key(len(context['excelcontent']['headers']))
    sheet.merge_cells("A1:" + letter + "1")

    csv_file.save(excel_filename)


def get_excel_default_key(index):
    if index < 0:
        return 'A'
    firstletter_index = index // 26
    secondletter_index = index % 26
    if index >= 0 and index <= 701:
        unicode = secondletter_index + 65
        secondletter = chr(unicode)
        firstletter = ''
        if firstletter_index == 0:
            firstletter = ''
        elif firstletter_index > 0:
            unicode = (firstletter_index - 1) + 65
            firstletter = chr(unicode)

        return firstletter + secondletter

    elif index > 701:
        return None

def set_width_adjustment(ws, type='', letter={}):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.coordinate in ws.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                # max length calc
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

            except:
                pass

        adjusted_width = max_length + 7
        if adjusted_width >= 30:
            adjusted_width = 30

        ws.column_dimensions[column].width = adjusted_width